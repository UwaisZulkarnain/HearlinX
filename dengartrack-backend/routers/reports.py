from datetime import datetime
from io import BytesIO
import uuid

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from auth.dependencies import coordinator_or_unhs, moh_only, require_role
from auth.models import (
    MonthlyReportSummary,
    NationalHospitalSummary,
    NationalSummaryOut,
    BenchmarkReport,
    CoverageReport,
    WardBreakdownReport,
    WardBreakdownItem,
)
from db.database import get_db

router = APIRouter(prefix="/reports", tags=["reports"])


def get_monthly_summary(db: Session, hospital_id: str, year: int, month: int):
    return db.execute(
        text(
            """
            SELECT
                h.id AS hospital_id,
                h.name AS hospital_name,
                COALESCE(sc.total_screenings, 0) AS total_screenings,
                COALESCE(sc.total_pass, 0) AS total_pass,
                COALESCE(sc.total_refer, 0) AS total_refer,
                COALESCE(sc.total_not_tested, 0) AS total_not_tested,
                COALESCE(fc.total_ltfu, 0) AS total_ltfu
            FROM hospitals h
            LEFT JOIN (
                SELECT
                    hospital_id,
                    COUNT(*) AS total_screenings,
                    COALESCE(SUM(CASE WHEN ear_left = 'pass' AND ear_right = 'pass' THEN 1 ELSE 0 END), 0) AS total_pass,
                    COALESCE(SUM(CASE WHEN ear_left = 'refer' OR ear_right = 'refer' THEN 1 ELSE 0 END), 0) AS total_refer,
                    COALESCE(SUM(CASE WHEN (ear_left = 'refer' OR ear_right = 'refer') THEN 0 WHEN (ear_left = 'not_tested' OR ear_right = 'not_tested') THEN 1 ELSE 0 END), 0) AS total_not_tested
                FROM screenings
                WHERE EXTRACT(YEAR FROM screening_date) = :year
                  AND EXTRACT(MONTH FROM screening_date) = :month
                GROUP BY hospital_id
            ) sc ON sc.hospital_id = h.id
            LEFT JOIN (
                SELECT hospital_id, COUNT(*) AS total_ltfu
                FROM follow_ups
                WHERE status = 'lost_to_followup'
                  AND EXTRACT(YEAR FROM updated_at) = :year
                  AND EXTRACT(MONTH FROM updated_at) = :month
                GROUP BY hospital_id
            ) fc ON fc.hospital_id = h.id
            WHERE h.id = :hospital_id
            """
        ),
        {"hospital_id": hospital_id, "year": year, "month": month},
    ).fetchone()


def get_all_hospitals_monthly_summary(db: Session, year: int, month: int):
    return db.execute(
        text(
            """
            SELECT
                COALESCE(sc.total_screenings, 0) AS total_screenings,
                COALESCE(sc.total_pass, 0) AS total_pass,
                COALESCE(sc.total_refer, 0) AS total_refer,
                COALESCE(sc.total_not_tested, 0) AS total_not_tested,
                COALESCE(fc.total_ltfu, 0) AS total_ltfu
            FROM (
                SELECT
                    COUNT(*) AS total_screenings,
                    COALESCE(SUM(CASE WHEN ear_left = 'pass' AND ear_right = 'pass' THEN 1 ELSE 0 END), 0) AS total_pass,
                    COALESCE(SUM(CASE WHEN ear_left = 'refer' OR ear_right = 'refer' THEN 1 ELSE 0 END), 0) AS total_refer,
                    COALESCE(SUM(CASE WHEN (ear_left = 'refer' OR ear_right = 'refer') THEN 0 WHEN (ear_left = 'not_tested' OR ear_right = 'not_tested') THEN 1 ELSE 0 END), 0) AS total_not_tested
                FROM screenings
                WHERE EXTRACT(YEAR FROM screening_date) = :year
                  AND EXTRACT(MONTH FROM screening_date) = :month
            ) sc
            CROSS JOIN (
                SELECT COUNT(*) AS total_ltfu
                FROM follow_ups
                WHERE status = 'lost_to_followup'
                  AND EXTRACT(YEAR FROM updated_at) = :year
                  AND EXTRACT(MONTH FROM updated_at) = :month
            ) fc
            """
        ),
        {"year": year, "month": month},
    ).fetchone()


@router.get("/monthly", response_model=MonthlyReportSummary, tags=["reports"])
def monthly_report(
    year: int | None = Query(default=None, ge=2000),
    month: int | None = Query(default=None, ge=1, le=12),
    current_user: dict = Depends(coordinator_or_unhs),
    db: Session = Depends(get_db),
):
    now = datetime.utcnow()
    target_year = year or now.year
    target_month = month or now.month

    if current_user["role"] == "unhs_coordinator":
        summary = get_all_hospitals_monthly_summary(db, target_year, target_month)
        return MonthlyReportSummary(
            hospital_id=uuid.UUID("00000000-0000-0000-0000-000000000000"),
            hospital_name="All Hospitals",
            year=target_year,
            month=target_month,
            total_screenings=summary.total_screenings or 0,
            total_pass=summary.total_pass or 0,
            total_refer=summary.total_refer or 0,
            total_not_tested=summary.total_not_tested or 0,
            total_ltfu=summary.total_ltfu or 0,
        )

    summary = get_monthly_summary(db, current_user["hospital_id"], target_year, target_month)
    return MonthlyReportSummary(
        hospital_id=summary.hospital_id,
        hospital_name=summary.hospital_name,
        year=target_year,
        month=target_month,
        total_screenings=summary.total_screenings or 0,
        total_pass=summary.total_pass or 0,
        total_refer=summary.total_refer or 0,
        total_not_tested=summary.total_not_tested or 0,
        total_ltfu=summary.total_ltfu or 0,
    )


@router.get("/export", response_model=None, tags=["reports"])
def export_report(
    year: int | None = Query(default=None, ge=2000),
    month: int | None = Query(default=None, ge=1, le=12),
    current_user: dict = Depends(coordinator_or_unhs),
    db: Session = Depends(get_db),
):
    now = datetime.utcnow()
    target_year = year or now.year
    target_month = month or now.month
    hospital_id = current_user.get("hospital_id")

    if current_user["role"] == "unhs_coordinator":
        rows = db.execute(
            text(
                """
                SELECT
                    h.name AS hospital_name,
                    COALESCE(sc.total_screenings, 0) AS total_screenings,
                    COALESCE(sc.total_pass, 0) AS total_pass,
                    COALESCE(sc.total_refer, 0) AS total_refer,
                    COALESCE(sc.total_not_tested, 0) AS total_not_tested,
                    COALESCE(fc.total_ltfu, 0) AS total_ltfu
                FROM hospitals h
                LEFT JOIN (
                    SELECT
                        hospital_id,
                        COUNT(*) AS total_screenings,
                        COALESCE(SUM(CASE WHEN ear_left = 'pass' AND ear_right = 'pass' THEN 1 ELSE 0 END), 0) AS total_pass,
                        COALESCE(SUM(CASE WHEN ear_left = 'refer' OR ear_right = 'refer' THEN 1 ELSE 0 END), 0) AS total_refer,
                        COALESCE(SUM(CASE WHEN (ear_left = 'refer' OR ear_right = 'refer') THEN 0 WHEN (ear_left = 'not_tested' OR ear_right = 'not_tested') THEN 1 ELSE 0 END), 0) AS total_not_tested
                    FROM screenings
                    WHERE EXTRACT(YEAR FROM screening_date) = :year
                      AND EXTRACT(MONTH FROM screening_date) = :month
                    GROUP BY hospital_id
                ) sc ON sc.hospital_id = h.id
                LEFT JOIN (
                    SELECT hospital_id, COUNT(*) AS total_ltfu
                    FROM follow_ups
                    WHERE status = 'lost_to_followup'
                      AND EXTRACT(YEAR FROM updated_at) = :year
                      AND EXTRACT(MONTH FROM updated_at) = :month
                    GROUP BY hospital_id
                ) fc ON fc.hospital_id = h.id
                ORDER BY h.name ASC
                """
            ),
            {"year": target_year, "month": target_month},
        ).fetchall()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "All Hospitals Summary"
        sheet.append(["Hospital", "Year", "Month", "Total Screenings", "Total Pass", "Total Refer", "Total Not Tested", "Total LTFU"])
        for row in rows:
            sheet.append([
                row.hospital_name,
                target_year,
                target_month,
                row.total_screenings or 0,
                row.total_pass or 0,
                row.total_refer or 0,
                row.total_not_tested or 0,
                row.total_ltfu or 0,
            ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"dengartrack-all-hospitals-summary-{target_year:04d}-{target_month:02d}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    summary = get_monthly_summary(db, hospital_id, target_year, target_month)
    detail_rows = db.execute(
        text(
            """
            SELECT
                b.system_id,
                s.screening_type,
                s.ear_left,
                s.ear_right,
                s.attempt_number,
                s.screening_date,
                u.full_name AS screener_name
            FROM screenings s
            JOIN babies b ON b.id = s.baby_id
            JOIN users u ON u.id = s.screener_id
            WHERE s.hospital_id = :hospital_id
              AND EXTRACT(YEAR FROM s.screening_date) = :year
              AND EXTRACT(MONTH FROM s.screening_date) = :month
            ORDER BY s.screening_date DESC
            """
        ),
        {"hospital_id": hospital_id, "year": target_year, "month": target_month},
    ).fetchall()

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Monthly Summary"
    summary_sheet.append(["Hospital", "Year", "Month", "Total Screenings", "Total Pass", "Total Refer", "Total Not Tested", "Total LTFU"])
    summary_sheet.append([
        summary.hospital_name,
        target_year,
        target_month,
        summary.total_screenings or 0,
        summary.total_pass or 0,
        summary.total_refer or 0,
        summary.total_not_tested or 0,
        summary.total_ltfu or 0,
    ])

    details_sheet = workbook.create_sheet("Screenings")
    details_sheet.append(["System ID", "Screening Type", "Left Ear", "Right Ear", "Attempt", "Screening Date", "Screener"])
    for row in detail_rows:
        details_sheet.append([
            row.system_id,
            row.screening_type,
            row.ear_left,
            row.ear_right,
            row.attempt_number,
            row.screening_date.isoformat() if row.screening_date else None,
            row.screener_name,
        ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"dengartrack-report-{target_year:04d}-{target_month:02d}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/national-summary", response_model=NationalSummaryOut, tags=["reports"])
def national_summary(
    year: int | None = Query(default=None, ge=2000),
    month: int | None = Query(default=None, ge=1, le=12),
    current_user: dict = Depends(moh_only),
    db: Session = Depends(get_db),
):
    now = datetime.utcnow()
    target_year = year or now.year
    target_month = month or now.month

    rows = db.execute(
        text(
            """
            SELECT
                h.id AS hospital_id,
                h.name AS hospital_name,
                COALESCE(sc.total_screenings, 0) AS total_screenings,
                COALESCE(sc.total_pass, 0) AS total_pass,
                COALESCE(sc.total_refer, 0) AS total_refer,
                COALESCE(sc.total_not_tested, 0) AS total_not_tested,
                COALESCE(fc.total_ltfu, 0) AS total_ltfu
            FROM hospitals h
            LEFT JOIN (
                SELECT
                    hospital_id,
                    COUNT(*) AS total_screenings,
                    COALESCE(SUM(CASE WHEN ear_left = 'pass' OR ear_right = 'pass' THEN 1 ELSE 0 END), 0) AS total_pass,
                    COALESCE(SUM(CASE WHEN ear_left = 'refer' OR ear_right = 'refer' THEN 1 ELSE 0 END), 0) AS total_refer,
                    COALESCE(SUM(CASE WHEN ear_left = 'not_tested' AND ear_right = 'not_tested' THEN 1 ELSE 0 END), 0) AS total_not_tested
                FROM screenings
                WHERE EXTRACT(YEAR FROM screening_date) = :year
                  AND EXTRACT(MONTH FROM screening_date) = :month
                GROUP BY hospital_id
            ) sc ON sc.hospital_id = h.id
            LEFT JOIN (
                SELECT hospital_id, COUNT(*) AS total_ltfu
                FROM follow_ups
                WHERE status = 'lost_to_followup'
                  AND EXTRACT(YEAR FROM updated_at) = :year
                  AND EXTRACT(MONTH FROM updated_at) = :month
                GROUP BY hospital_id
            ) fc ON fc.hospital_id = h.id
            ORDER BY h.name ASC
            """
        ),
        {"year": target_year, "month": target_month},
    ).fetchall()

    hospitals = [
        NationalHospitalSummary(
            hospital_id=row.hospital_id,
            hospital_name=row.hospital_name,
            total_screenings=row.total_screenings or 0,
            total_pass=row.total_pass or 0,
            total_refer=row.total_refer or 0,
            total_not_tested=row.total_not_tested or 0,
            total_ltfu=row.total_ltfu or 0,
        )
        for row in rows
    ]

    return NationalSummaryOut(
        year=target_year,
        month=target_month,
        total_hospitals=len(hospitals),
        total_screenings=sum(item.total_screenings for item in hospitals),
        total_pass=sum(item.total_pass for item in hospitals),
        total_refer=sum(item.total_refer for item in hospitals),
        total_not_tested=sum(item.total_not_tested for item in hospitals),
        total_ltfu=sum(item.total_ltfu for item in hospitals),
        hospitals=hospitals,
    )


@router.get("/benchmark", response_model=BenchmarkReport, tags=["reports"])
def benchmark_report(
    current_user: dict = Depends(coordinator_or_unhs),
    db: Session = Depends(get_db),
):
    """
    Returns 1-3-6 KKM compliance metrics for the coordinator's hospital:
    - screened_by_1_month: babies screened within 30 days of date_of_birth
    - diagnosed_by_3_months: follow_ups with appointment_date within 90 days of date_of_birth
    - total_eligible: total babies registered in hospital
    - Scoped by coordinator's hospital_id from JWT
    """
    hospital_id = current_user.get("hospital_id")

    result = db.execute(
        text(
            """
            SELECT
                COALESCE(SUM(CASE 
                    WHEN AGE(s.screening_date::date, b.date_of_birth) <= INTERVAL '30 days'
                    THEN 1 ELSE 0 
                END), 0) AS screened_by_1_month,
                COALESCE(SUM(CASE 
                    WHEN f.appointment_date IS NOT NULL 
                    AND AGE(f.appointment_date::date, b.date_of_birth) <= INTERVAL '90 days'
                    THEN 1 ELSE 0 
                END), 0) AS diagnosed_by_3_months
            FROM babies b
            LEFT JOIN screenings s ON s.baby_id = b.id
            LEFT JOIN follow_ups f ON f.baby_id = b.id
            WHERE b.hospital_id = :hospital_id
            """
        ),
        {"hospital_id": hospital_id},
    ).fetchone()

    total_eligible = db.execute(
        text("SELECT COUNT(*) AS total FROM babies WHERE hospital_id = :hospital_id"),
        {"hospital_id": hospital_id},
    ).scalar()

    screened_by_1_month = result.screened_by_1_month or 0
    diagnosed_by_3_months = result.diagnosed_by_3_months or 0
    total_eligible = total_eligible or 0

    screened_by_1_month_pct = (
        (screened_by_1_month / total_eligible * 100) if total_eligible > 0 else 0
    )
    diagnosed_by_3_months_pct = (
        (diagnosed_by_3_months / total_eligible * 100) if total_eligible > 0 else 0
    )

    return BenchmarkReport(
        screened_by_1_month=screened_by_1_month,
        diagnosed_by_3_months=diagnosed_by_3_months,
        total_eligible=total_eligible,
        screened_by_1_month_pct=round(screened_by_1_month_pct, 2),
        diagnosed_by_3_months_pct=round(diagnosed_by_3_months_pct, 2),
    )


@router.get("/coverage", response_model=CoverageReport, tags=["reports"])
def coverage_report(
    current_user: dict = Depends(require_role("coordinator", "unhs_coordinator", "moh")),
    db: Session = Depends(get_db),
):
    """
    Returns coverage rate for coordinator's hospital:
    - total_babies_registered: count of babies in hospital
    - total_babies_screened: count of distinct babies with at least one screening
    - coverage_rate_pct: (screened / registered) * 100
    Scoped by hospital_id from JWT.
    Accessible by coordinator, unhs_coordinator, moh roles.
    """
    hospital_id = current_user.get("hospital_id")

    total_registered = db.execute(
        text("SELECT COUNT(*) AS total FROM babies WHERE hospital_id = :hospital_id"),
        {"hospital_id": hospital_id},
    ).scalar()

    total_screened = db.execute(
        text(
            """
            SELECT COUNT(DISTINCT baby_id) AS total 
            FROM screenings 
            WHERE hospital_id = :hospital_id
            """
        ),
        {"hospital_id": hospital_id},
    ).scalar()

    total_registered = total_registered or 0
    total_screened = total_screened or 0

    coverage_rate_pct = (
        (total_screened / total_registered * 100) if total_registered > 0 else 0
    )

    return CoverageReport(
        total_babies_registered=total_registered,
        total_babies_screened=total_screened,
        coverage_rate_pct=round(coverage_rate_pct, 2),
    )


@router.get("/ward-breakdown", response_model=WardBreakdownReport, tags=["reports"])
def ward_breakdown_report(
    current_user: dict = Depends(coordinator_or_unhs),
    db: Session = Depends(get_db),
):
    """
    Returns per-ward stats for coordinator's hospital:
    - ward: ward name
    - total_screenings: count of screenings in that ward
    - total_refer: screenings where ear_left='refer' OR ear_right='refer'
    - refer_rate_pct: (refer / total) * 100
    Scoped by hospital_id from JWT.
    """
    hospital_id = current_user.get("hospital_id")

    rows = db.execute(
        text(
            """
            SELECT
                b.ward,
                COUNT(*) AS total_screenings,
                COALESCE(SUM(CASE 
                    WHEN s.ear_left = 'refer' OR s.ear_right = 'refer' 
                    THEN 1 ELSE 0 
                END), 0) AS total_refer
            FROM screenings s
            JOIN babies b ON b.id = s.baby_id
            WHERE b.hospital_id = :hospital_id
            GROUP BY b.ward
            ORDER BY b.ward ASC NULLS LAST
            """
        ),
        {"hospital_id": hospital_id},
    ).fetchall()

    wards = []
    for row in rows:
        total_screenings = row.total_screenings or 0
        total_refer = row.total_refer or 0
        refer_rate_pct = (
            (total_refer / total_screenings * 100) if total_screenings > 0 else 0
        )

        wards.append(
            WardBreakdownItem(
                ward=row.ward,
                total_screenings=total_screenings,
                total_refer=total_refer,
                refer_rate_pct=round(refer_rate_pct, 2),
            )
        )

    return WardBreakdownReport(wards=wards)
